# -*- coding: utf-8 -*-
"""
verify.py — pAIM1 scFv QC 웹도구 자체 검사
==============================================================================
    python verify.py

검사별 통과/실패를 표로 출력하고, 하나라도 실패하면 종료코드 1 을 반환합니다.

  [A] 구문   core.py / xlsx_writer.py / verify.py 의 AST, index.html 의 JS 구문,
             % 포맷 자리표시자 대 인자 개수, 함수 안의 미정의 이름
  [B] 계약   index.html 이 참조하는 컬럼명·키가 core.py 가 실제로 주는 것인가
  [C] 누출   index.html 에 서열이 하드코딩되어 있지 않은가
  [D] 회귀   testdata/ 로 core.analyze 를 돌려 알려진 값이 재현되는가
  [E] 단위   합성 서열로 core.check_landmark 의 5 개 상태를 직접 확인하고,
             index.html 의 badge() 가 그 표시값을 모두 처리하는지 대조
  [F] GLUE    index.html 의 GLUE 를 꺼내 실행해 배치별 호출과 병합을 검증
  [G] 대조군  testdata/negctrl/ 실측 .ab1 로 대조군 판정과 인서트 지문을 고정

RULES 커버리지
------------------------------------------------------------------------------
core.RULES 8 개를 하나씩 교란해 실측한 결과입니다. RULES 를 건드릴 때
무엇이 보호되고 무엇이 보호되지 않는지 여기서 확인하세요.

  키                  감지하는 검사              근거 / 공백 사유
  ------------------  -------------------------  --------------------------------
  AL_MATCH            D 클론별 판정·길이         1->2 에서 c01 verdict 가
                      E7, E11                    NO_LINKER -> LINKER_DEL 로 뒤집힘
  AL_MIS              D 클론별 판정·길이         -1->-2 에서 위와 동일
                      E7, E11
  AL_GAP              E9, E10                    E9 는 갭 1 개와 치환 2 개가
                                                 박빙인 결실 위치 3 을 씀
  AL_FLANK            E10                        삽입이 DP 창을 넘어설 때만
                                                 갈림. 결실로는 갈리지 않음
  FLAG_SEV            D 클론별 판정·길이 (일부)  testdata 가 실제로 만드는 7 개
                                                 플래그만. NO_LINKER 3->1 은 잡히나
                                                 MIXED 등 미발생 플래그는 미시험
  REPEAT_MAX_PERIOD   미시험                     60->10 이면 c03 의 TANDEM_REPEAT
                                                 가 사라지지만, 플래그 목록을
                                                 단언하는 검사가 없음
  EXO_TRIM            미시험                     0~6 어느 값에서도 testdata 의
                                                 CDR3 경계 판정이 바뀌지 않음
                                                 (민감한 클론이 없음)
  OVERLONG_ZONE       미시험                     5->200 이면 overlong_suspect 에
                                                 c01, c03 이 생기지만 이를
                                                 단언하는 검사가 없음

표준 라이브러리만 사용합니다.
node 와 openpyxl 은 있으면 쓰고 없으면 해당 검사를 "건너뜀" 으로 표시합니다.
기대값은 별도 파일이 아니라 이 파일 안의 EXPECT 리터럴에 둡니다
(.gitignore 가 *.csv 를 막으므로).
"""

import ast
import builtins
import io
import os
import re
import shutil
import json
import string
import subprocess
import sys
import tempfile
import traceback
import unicodedata

VERIFY_VERSION = "2.4"

ROOT = os.path.dirname(os.path.abspath(__file__))
PY_FILES = ("core.py", "xlsx_writer.py", "verify.py")
HTML_FILE = "index.html"
TESTDATA = "testdata"

# --- [B] DESIGN_DOC 누락 허용 목록 --------------------------------------------
# 이슈 3 해결로 비었습니다. DESIGN_DEFAULTS 와 DESIGN_DOC 의 키는 이제 완전히
# 일치해야 하며, 하나라도 어긋나면 실패합니다.
KNOWN_DESIGN_DOC_GAP = set()

# --- [C] 서열이 아니라 문자 클래스이므로 제외 --------------------------------
MARKSEQ_REGEX = r"/\b[ACGT]{8,}\b/g"

# --- [D] 기대값 --------------------------------------------------------------
# [D] 회귀는 assigned 모드로 돌립니다. 배치 지정 대조까지 함께 지나가게 하려는
# 것이며, 모드 값은 core 의 상수를 참조하지 않고 리터럴로 둡니다. 상수 값이 바뀌면
# design_hash 가 달라져 [D] 가 잡아야 하기 때문입니다.
REG_OVERRIDES = {"analysis_mode": "assigned",
                 "batch_vh_family": "VH6", "batch_chain": "kappa"}
REG_META = {"batch_label": "VH6-VK", "batch_date": "260819"}

_STEM = "260819_VH6-VK_c%s_pAIM1-seq-For"      # 클론 ID = 확장자 뗀 파일명

EXPECT = {
    "param_hash": "3473927a",
    "design_hash": "8b1eab32",
    "clones": [
        {"id": _STEM % "01", "tag": "c01", "verdict": "NO_LINKER",  "insert": 276, "mix": 2.5},
        {"id": _STEM % "02", "tag": "c02", "verdict": "PASS",       "insert": 728,
         "d1": 366, "d2": 309, "mix": 2.1},
        {"id": _STEM % "03", "tag": "c03", "verdict": "FRAMESHIFT", "insert": 735,
         "d1": 346, "d2": 336, "mix": 3.9},
        {"id": _STEM % "04", "tag": "c04", "verdict": "PASS",       "insert": 746,
         "d1": 372, "d2": 321, "mix": 1.9},
    ],
    "cdr3_median": 10,
    "vl": [("IGKV4", 1), ("IGKV3", 1)],
    "jh": [("JH1|JH2", 1), ("JH4", 1)],
    "fasta_n": 2,
}

# JS 내장 프로퍼티 — 데이터 키가 아니므로 하위키 대조에서 제외합니다.
JS_PROPS = {
    "length", "forEach", "map", "filter", "find", "join", "split", "slice",
    "replace", "indexOf", "push", "pop", "concat", "sort", "reverse", "some",
    "every", "includes", "startsWith", "endsWith", "toFixed", "toString",
    "trim", "keys", "values", "entries", "has", "add", "delete", "get", "set",
    "size", "then", "catch", "flat", "reduce", "charAt", "substring", "padStart",
}


# =============================================================================
#  결과 수집과 표 출력
# =============================================================================
PASS, FAIL, SKIP = "PASS", "FAIL", "SKIP"
_MARK = {PASS: "통과", FAIL: "실패", SKIP: "건너뜀"}


class Report(object):
    def __init__(self):
        self.rows = []

    def add(self, section, name, status, detail=""):
        self.rows.append((section, name, status, detail))

    def ok(self, section, name, cond, detail_ok="", detail_ng=""):
        if cond:
            self.add(section, name, PASS, detail_ok)
        else:
            self.add(section, name, FAIL, detail_ng or detail_ok)

    def count(self, section, status):
        return sum(1 for r in self.rows if r[0] == section and r[2] == status)

    def failed(self):
        return any(r[2] == FAIL for r in self.rows)


def guard(report, section, label, fn, *args):
    """검사 하나가 예외로 죽어도 표를 끝까지 낸다. 예외는 그 자체로 실패."""
    try:
        fn(*args)
    except Exception as e:
        tb = traceback.extract_tb(sys.exc_info()[2])
        where = ""
        if tb:
            last = tb[-1]
            where = " @ %s:%d" % (os.path.basename(last.filename), last.lineno)
        report.add(section, label, FAIL,
                   "검사 중 예외 %s: %s%s" % (type(e).__name__, e, where))


def width(text):
    """한글을 2칸으로 세는 표시폭."""
    n = 0
    for ch in str(text):
        n += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
    return n


def pad(text, w):
    text = str(text)
    return text + " " * max(0, w - width(text))


def clip(text, w):
    text = str(text).replace("\n", " ")
    if width(text) <= w:
        return text
    out = ""
    for ch in text:
        if width(out) + width(ch) > w - 1:
            return out + "…"
        out += ch
    return out


def print_table(report):
    head = ("구분", "검사", "결과", "상세")
    detail_w = 62
    body = [(r[0], r[1], _MARK[r[2]], clip(r[3], detail_w)) for r in report.rows]
    w0 = max([width(head[0])] + [width(r[0]) for r in body])
    w1 = max([width(head[1])] + [width(r[1]) for r in body])
    w2 = max([width(head[2])] + [width(r[2]) for r in body])
    w3 = max([width(head[3])] + [width(r[3]) for r in body])
    line = "  ".join(("-" * w0, "-" * w1, "-" * w2, "-" * w3))
    print(line)
    print("  ".join((pad(head[0], w0), pad(head[1], w1), pad(head[2], w2), head[3])))
    print(line)
    for r in body:
        print("  ".join((pad(r[0], w0), pad(r[1], w1), pad(r[2], w2), r[3])))
    print(line)


# =============================================================================
#  파일 읽기
# =============================================================================
def read_text(name):
    with io.open(os.path.join(ROOT, name), encoding="utf-8") as fh:
        return fh.read()


def parse_py(name):
    return ast.parse(read_text(name), filename=name)


# =============================================================================
#  index.html 조각 내기
# =============================================================================
_SCRIPT_RE = re.compile(r"<script\b([^>]*)>(.*?)</script>", re.S | re.I)


def script_blocks(html):
    """외부 src 가 아닌 <script> 블록의 본문 목록."""
    out = []
    for m in _SCRIPT_RE.finditer(html):
        if "src=" in m.group(1).lower():
            continue
        out.append(m.group(2))
    return out


def glue_source(js):
    """index.html 안의 GLUE 템플릿 리터럴(파이썬 접착부) 본문."""
    m = re.search(r"const\s+GLUE\s*=\s*`", js)
    if not m:
        return None
    start = m.end()
    end = js.find("`", start)
    if end < 0:
        return None
    return js[start:end]


_JS_STR_RE = re.compile(r"'((?:[^'\\\n]|\\.)*)'|\"((?:[^\"\\\n]|\\.)*)\"")


def js_strings(js):
    out = []
    for m in _JS_STR_RE.finditer(js):
        out.append(m.group(1) if m.group(1) is not None else m.group(2))
    return out


def js_array_literal(js, decl):
    """`const NAME = [...]` / `const NAME = new Set([...])` 의 문자열 항목."""
    m = re.search(re.escape(decl) + r"\s*=\s*(?:new\s+Set\s*\()?\s*\[", js)
    if not m:
        return None
    start = js.index("[", m.end() - 1)
    depth = 0
    for j in range(start, len(js)):
        if js[j] == "[":
            depth += 1
        elif js[j] == "]":
            depth -= 1
            if depth == 0:
                return js_strings(js[start:j + 1])
    return None


def js_function_body(js, name):
    """function NAME(...) { ... } 의 본문. 못 찾으면 None."""
    m = re.search(r"function\s+" + re.escape(name) + r"\s*\([^)]*\)\s*\{", js)
    if not m:
        return None
    start = m.end() - 1
    depth = 0
    for j in range(start, len(js)):
        if js[j] == "{":
            depth += 1
        elif js[j] == "}":
            depth -= 1
            if depth == 0:
                return js[start:j + 1]
    return None


def js_member_refs(js, obj):
    """`obj.key` 와 `obj.key.sub` 참조를 (key, sub) 쌍으로 모은다."""
    pat = re.compile(r"(?<![A-Za-z0-9_$.])" + re.escape(obj) +
                     r"\.([A-Za-z_$][A-Za-z0-9_$]*)(?:\.([A-Za-z_$][A-Za-z0-9_$]*))?")
    out = []
    for m in pat.finditer(js):
        out.append((m.group(1), m.group(2)))
    return out


# =============================================================================
#  [A] 구문
# =============================================================================
def check_ast(report):
    for name in PY_FILES:
        try:
            parse_py(name)
            report.add("A", name + " AST 파싱", PASS, "구문 오류 없음")
        except SyntaxError as e:
            report.add("A", name + " AST 파싱", FAIL,
                       "%s:%s %s" % (name, e.lineno, e.msg))


def node_error_line(stderr, tmpdir):
    """node --check 출력에서 사람이 읽을 한 줄만 뽑는다 (임시경로 제거)."""
    lines = [x.strip() for x in (stderr or "").splitlines() if x.strip()]
    pick = ""
    for x in lines:
        if re.match(r"^\w*Error\b", x):
            pick = x
            break
    if not pick:
        pick = lines[0] if lines else "node --check 실패"
    for x in lines:
        m = re.search(r"\.js:(\d+)", x)
        if m:
            pick = "블록 %s행 · %s" % (m.group(1), pick)
            break
    return pick.replace(tmpdir, "").replace(os.sep + "block", "block")


def check_js_syntax(report, html):
    node = shutil.which("node")
    blocks = script_blocks(html)
    if not blocks:
        report.add("A", "index.html JS 구문", FAIL, "<script> 블록을 찾지 못했습니다")
        return
    if not node:
        report.add("A", "index.html JS 구문", SKIP,
                   "node 없음 · <script> %d 블록 미검사" % len(blocks))
        return
    tmp = tempfile.mkdtemp(prefix="scfvqc_verify_")
    try:
        bad = []
        for i, src in enumerate(blocks):
            path = os.path.join(tmp, "block%d.js" % (i + 1))
            with io.open(path, "w", encoding="utf-8") as fh:
                fh.write(src)
            p = subprocess.run([node, "--check", path], capture_output=True,
                               text=True, encoding="utf-8", errors="replace")
            if p.returncode != 0:
                bad.append("블록%d: %s" % (i + 1, node_error_line(p.stderr, tmp)))
        report.ok("A", "index.html JS 구문", not bad,
                  "node --check 통과 · <script> %d 블록" % len(blocks),
                  " / ".join(bad))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


_PCT_RE = re.compile(
    r"%(%|[-+ #0]*(?:\*|\d+)?(?:\.(?:\*|\d+))?[hlL]?[diouxXeEfFgGcrsa])")

# 우변이 이 노드면 값 개수를 정적으로 알 수 없으므로 검사하지 않습니다.
UNKNOWN_RHS = (ast.Name, ast.Call, ast.Subscript, ast.Attribute,
               ast.GeneratorExp, ast.Dict, ast.DictComp, ast.Starred)


def pct_placeholders(text):
    """%% 를 제외한 자리표시자 개수. 개수를 알 수 없으면 None."""
    if "%(" in text:
        return None
    n = 0
    for m in _PCT_RE.finditer(text):
        if m.group(1) == "%":
            continue
        if "*" in m.group(1):
            return None
        n += 1
    return n


def rhs_arity(node):
    """% 우변이 공급하는 값의 개수. 알 수 없으면 None."""
    if isinstance(node, ast.Tuple):
        if any(isinstance(e, ast.Starred) for e in node.elts):
            return None
        return len(node.elts)
    if isinstance(node, UNKNOWN_RHS):
        return None
    return 1


def format_field_arity(text):
    """.format 의 위치 자리표시자 개수. 이름/속성 참조가 섞이면 None."""
    auto, indexed = 0, -1
    for _lit, field, _spec, _conv in string.Formatter().parse(text):
        if field is None:
            continue
        head = field.split(".")[0].split("[")[0]
        if head == "":
            auto += 1
        elif head.isdigit():
            indexed = max(indexed, int(head))
        else:
            return None
    if auto and indexed >= 0:
        return None
    return auto if auto else (indexed + 1)


def check_format_arity(report):
    bad, checked, skipped = [], 0, 0
    for name in PY_FILES:
        tree = parse_py(name)
        for node in ast.walk(tree):
            if (isinstance(node, ast.BinOp) and isinstance(node.op, ast.Mod)
                    and isinstance(node.left, ast.Constant)
                    and isinstance(node.left.value, str)):
                need = pct_placeholders(node.left.value)
                have = rhs_arity(node.right)
                if need is None or have is None:
                    skipped += 1
                    continue
                checked += 1
                if need != have:
                    bad.append("%s:%d %% 자리 %d 대 인자 %d"
                               % (name, node.lineno, need, have))
            if (isinstance(node, ast.Call)
                    and isinstance(node.func, ast.Attribute)
                    and node.func.attr == "format"
                    and isinstance(node.func.value, ast.Constant)
                    and isinstance(node.func.value.value, str)):
                if node.keywords or any(isinstance(a, ast.Starred) for a in node.args):
                    skipped += 1
                    continue
                need = format_field_arity(node.func.value.value)
                if need is None:
                    skipped += 1
                    continue
                checked += 1
                if need != len(node.args):
                    bad.append("%s:%d .format 자리 %d 대 인자 %d"
                               % (name, node.lineno, need, len(node.args)))
    report.ok("A", "포맷 자리표시자 대조", not bad,
              "대조 %d 건 · 개수 불명으로 제외 %d 건" % (checked, skipped),
              " / ".join(bad[:4]) + (" 외 %d" % (len(bad) - 4) if len(bad) > 4 else ""))


def _bind_targets(node, out):
    if isinstance(node, ast.Name):
        out.add(node.id)
    elif isinstance(node, (ast.Tuple, ast.List)):
        for e in node.elts:
            _bind_targets(e, out)
    elif isinstance(node, ast.Starred):
        _bind_targets(node.value, out)


def arg_names(args):
    out = set()
    for a in list(args.posonlyargs) + list(args.args) + list(args.kwonlyargs):
        out.add(a.arg)
    if args.vararg:
        out.add(args.vararg.arg)
    if args.kwarg:
        out.add(args.kwarg.arg)
    return out


def bound_names(body):
    """한 스코프가 묶는 이름 전부. 중첩 함수/람다 안으로는 들어가지 않는다."""
    out = set()
    stack = list(body)
    while stack:
        n = stack.pop()
        if isinstance(n, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            out.add(n.name)
            continue
        if isinstance(n, ast.Lambda):
            continue
        if isinstance(n, ast.Assign):
            for t in n.targets:
                _bind_targets(t, out)
        elif isinstance(n, (ast.AugAssign, ast.AnnAssign, ast.NamedExpr)):
            _bind_targets(n.target, out)
        elif isinstance(n, (ast.For, ast.AsyncFor)):
            _bind_targets(n.target, out)
        elif isinstance(n, (ast.With, ast.AsyncWith)):
            for it in n.items:
                if it.optional_vars is not None:
                    _bind_targets(it.optional_vars, out)
        elif isinstance(n, ast.ExceptHandler):
            if n.name:
                out.add(n.name)
        elif isinstance(n, (ast.Import, ast.ImportFrom)):
            for a in n.names:
                out.add((a.asname or a.name).split(".")[0])
        elif isinstance(n, (ast.Global, ast.Nonlocal)):
            out.update(n.names)
        elif isinstance(n, (ast.ListComp, ast.SetComp, ast.DictComp, ast.GeneratorExp)):
            for g in n.generators:
                _bind_targets(g.target, out)
        stack.extend(ast.iter_child_nodes(n))
    return out


def _scan_names(node, scope, undef, modname):
    """Load 로 쓰이는 이름이 스코프에 없으면 undef 에 담는다."""
    if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
        for d in node.decorator_list:
            _scan_names(d, scope, undef, modname)
        for d in list(node.args.defaults) + [x for x in node.args.kw_defaults if x]:
            _scan_names(d, scope, undef, modname)
        # 중첩함수는 바깥 스코프를 그대로 물려받는다 (클로저)
        inner = set(scope) | arg_names(node.args) | bound_names(node.body)
        for st in node.body:
            _scan_names(st, inner, undef, modname)
        return
    if isinstance(node, ast.Lambda):
        for d in list(node.args.defaults) + [x for x in node.args.kw_defaults if x]:
            _scan_names(d, scope, undef, modname)
        _scan_names(node.body, set(scope) | arg_names(node.args), undef, modname)
        return
    if isinstance(node, ast.ClassDef):
        for d in node.decorator_list + list(node.bases):
            _scan_names(d, scope, undef, modname)
        inner = set(scope) | bound_names(node.body)
        for st in node.body:
            _scan_names(st, inner, undef, modname)
        return
    if isinstance(node, (ast.ListComp, ast.SetComp, ast.GeneratorExp, ast.DictComp)):
        inner = set(scope)
        for g in node.generators:
            _bind_targets(g.target, inner)
        for g in node.generators:
            _scan_names(g.iter, inner, undef, modname)
            for c in g.ifs:
                _scan_names(c, inner, undef, modname)
        if isinstance(node, ast.DictComp):
            _scan_names(node.key, inner, undef, modname)
            _scan_names(node.value, inner, undef, modname)
        else:
            _scan_names(node.elt, inner, undef, modname)
        return
    if isinstance(node, ast.ExceptHandler):
        inner = set(scope)
        if node.name:
            inner.add(node.name)
        if node.type is not None:
            _scan_names(node.type, inner, undef, modname)
        for st in node.body:
            _scan_names(st, inner, undef, modname)
        return
    if isinstance(node, ast.Name):
        if isinstance(node.ctx, ast.Load) and node.id not in scope:
            undef.append("%s:%d %s" % (modname, node.lineno, node.id))
        return
    for c in ast.iter_child_nodes(node):
        _scan_names(c, scope, undef, modname)


def check_undefined_names(report):
    base = set(dir(builtins)) | {"__name__", "__file__", "__doc__", "__spec__"}
    undef = []
    for name in PY_FILES:
        tree = parse_py(name)
        scope = base | bound_names(tree.body)
        for st in tree.body:
            _scan_names(st, scope, undef, name)
    report.ok("A", "함수 안 미정의 이름", not undef,
              "검사 %d 파일 · 미정의 0 건" % len(PY_FILES),
              " / ".join(undef[:5]) +
              (" 외 %d" % (len(undef) - 5) if len(undef) > 5 else ""))


# =============================================================================
#  [B] 계약
# =============================================================================
def _dict_literal_keys(d):
    out = {}
    for k, v in zip(d.keys, d.values):
        if isinstance(k, ast.Constant) and isinstance(k.value, str):
            out[k.value] = v
    return out


def _returned_dict(rv):
    if isinstance(rv, ast.Dict):
        return rv
    if isinstance(rv, ast.Call) and rv.args and isinstance(rv.args[0], ast.Dict):
        return rv.args[0]      # json.dumps({...}) 형태
    return None


def provided_keys(fn):
    """함수가 돌려주는 dict 의 키 집합과, 값이 dict 리터럴인 키의 하위키 맵."""
    keys = {}
    names = set()
    for node in ast.walk(fn):
        if isinstance(node, ast.Return) and node.value is not None:
            d = _returned_dict(node.value)
            if d is not None:
                keys.update(_dict_literal_keys(d))
            elif isinstance(node.value, ast.Name):
                names.add(node.value.id)
    for node in ast.walk(fn):
        if not isinstance(node, ast.Assign):
            continue
        for t in node.targets:
            if isinstance(t, ast.Name) and t.id in names and isinstance(node.value, ast.Dict):
                keys.update(_dict_literal_keys(node.value))
            if (isinstance(t, ast.Subscript) and isinstance(t.value, ast.Name)
                    and t.value.id in names and isinstance(t.slice, ast.Constant)
                    and isinstance(t.slice.value, str)):
                keys.setdefault(t.slice.value, node.value)
    sub = dict((k, set(_dict_literal_keys(v)))
               for k, v in keys.items() if isinstance(v, ast.Dict))
    return set(keys), sub


def find_function(tree, name):
    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name == name:
            return node
    return None


def check_summary_headers(report, js, core):
    headers = set(core.SUMMARY_HEADERS)
    total, missing = 0, []
    for decl in ("const COMPACT", "const NUMCOL", "const MONOCOL", "const BADGECOL"):
        items = js_array_literal(js, decl)
        label = decl.split()[-1]
        if items is None:
            report.add("B", label + " 대 SUMMARY_HEADERS", FAIL,
                       "index.html 에서 " + decl + " 선언을 찾지 못했습니다")
            continue
        total += len(items)
        gone = [x for x in items if x not in headers]
        missing += gone
        report.ok("B", label + " 대 SUMMARY_HEADERS", not gone,
                  "%d 항목 전부 존재" % len(items),
                  "SUMMARY_HEADERS 에 없음: " + ", ".join(gone))
    if total:
        report.add("B", "컬럼명 복제 합계", PASS if not missing else FAIL,
                   "대조 %d 항목 · 누락 %d" % (total, len(missing)))


def analyze_keyset(report, core, core_tree, reg_out, reg_note):
    """[B] 가 대조 기준으로 삼을 analyze / compose 의 키 집합.

    analyze 를 실제로 돌렸으면 그 결과의 키를 그대로 쓴다. analyze 에는 오류
    경로와 성공 경로의 return 이 따로 있어 AST 추정은 엉뚱한 쪽을 읽을 수 있다.
    testdata 가 없어 돌리지 못했을 때만 AST 로 폴백하고 그 사실을 표시한다.
    """
    try:
        cfg = core.build_config(None, [])
    except Exception:
        cfg = {}
    if reg_out is not None:
        out_sub = dict((k, set(v)) for k, v in reg_out.items() if isinstance(v, dict))
        comp = reg_out.get("composition") or {}
        return set(reg_out), out_sub, "실행 결과", set(comp), "실행 결과"

    fn = find_function(core_tree, "analyze")
    out_keys, out_sub = provided_keys(fn) if fn else (set(), {})
    out_sub = dict(out_sub)
    try:
        comp = core.compose([], cfg)
    except Exception:
        comp = {}
    out_sub["config"] = set(cfg)
    out_sub["composition"] = set(comp)
    if not out_keys:
        report.add("B", "core.analyze 반환 키 추출", FAIL,
                   "analyze 의 반환 dict 를 AST 로도 읽지 못했습니다")
    return (out_keys, out_sub, "AST 폴백 · " + (reg_note or "실행 결과 없음"),
            set(comp), "빈 입력 실행")


def check_js_keys(report, js, core, keyset, doc_keys, doc_sub):
    out_keys, out_sub, out_src, comp, comp_src = keyset
    scan = core.scan_inputs([], "")
    scan_sub = dict((k, set(v)) for k, v in scan.items() if isinstance(v, dict))

    targets = [
        # analyze 반환을 가리키는 별칭들. RUN.merged 는 M 으로, 배치별 결과는
        # b.result 로 받으므로 두 경로 모두 analyze 의 키로 대조합니다.
        ("OUT", "core.analyze", set(out_keys), out_sub, out_src),
        ("M", "core.analyze (RUN.merged)", set(out_keys), out_sub, out_src),
        ("SCAN", "core.scan_inputs", set(scan), scan_sub, "실행 결과"),
        ("DOC", "js_docs (GLUE)", set(doc_keys), doc_sub, "AST"),
        ("C", "core.compose", set(comp), {}, comp_src),
    ]
    for obj, source, keys, sub, origin in targets:
        refs = js_member_refs(js, obj)
        if obj == "C" and not refs:
            report.add("B", "C.* 대 " + source, FAIL,
                       "composition 별칭 참조를 찾지 못했습니다")
            continue
        bad = []
        n = 0
        for key, subkey in refs:
            n += 1
            if key not in keys:
                bad.append("%s.%s" % (obj, key))
                continue
            if subkey and subkey not in JS_PROPS and key in sub and sub[key]:
                if subkey not in sub[key]:
                    bad.append("%s.%s.%s" % (obj, key, subkey))
        bad = sorted(set(bad))
        report.ok("B", obj + ".* 대 " + source, not bad,
                  "참조 %d 건 전부 존재 · 제공 키 %d (%s)" % (n, len(keys), origin),
                  "제공되지 않는 키: " + ", ".join(bad) + " (%s)" % origin)

    # 배치별 결과 : b.result.<키> 형태로 analyze 반환을 참조합니다.
    hits = re.findall(r"\.result\.([A-Za-z_$][A-Za-z0-9_$]*)", js)
    bad = sorted(set(k for k in hits if k not in out_keys and k not in JS_PROPS))
    report.ok("B", ".result.* 대 core.analyze", not bad,
              "참조 %d 건 전부 존재 (%s)" % (len(hits), out_src),
              "제공되지 않는 키: " + ", ".join(bad))

    # fillBatchOptions 의 guess 인자 (SCAN.guess 하위키)
    body = js_function_body(js, "fillBatchOptions")
    if body is not None and isinstance(scan.get("guess"), dict):
        used = sorted(set(k for k, _ in js_member_refs(body, "guess")))
        gone = [k for k in used if k not in scan["guess"]]
        report.ok("B", "guess.* 대 scan_inputs['guess']", not gone,
                  "참조 %s 전부 존재" % ", ".join(used) if used else "참조 없음",
                  "없는 키: " + ", ".join(gone))


# 배치 모드에서만 GLUE 의 js_analyze_batches 가 카드 값으로 덮어쓰는 키.
# 나머지는 readConfig 가 반드시 보내야 합니다 — 라이브러리 모드(js_analyze)는
# readConfig 의 cfg 를 그대로 core.analyze 에 넘기므로, 거기서 빠진 키는 조용히
# 기본값으로 돌아갑니다. 근거 구간을 GLUE 까지 넓히면 이 구멍을 못 잡습니다.
GLUE_SUPPLIED_KEYS = ("batch_vh_family", "batch_chain")


def check_readconfig_covers_design(report, js, core, glue=""):
    """설계 키가 index.html 을 통해 core 까지 전달되는가."""
    need = [k for k in core.DESIGN_KEYS if k not in GLUE_SUPPLIED_KEYS]
    parts, source = [], []
    for fname in ("buildDesignForm", "readConfig"):
        body = js_function_body(js, fname)
        if body is None:
            continue
        parts.append(body)
        source.append(fname)
    if not parts:
        parts, source = [js], ["script 전체 (함수 추출 실패)"]
    literal = set(js_strings("\n".join(parts)))
    dynamic = set()
    # cfg 를 채우는 것은 readConfig 이므로, 동적 생성 인정은 readConfig 가 직접
    # DOC.design 을 순회할 때만 합니다. buildDesignForm 이 조회 테이블을 만드느라
    # DOC.design 을 읽는 것만으로는 그 키가 cfg 에 실린다는 보장이 없습니다.
    if re.search(r"DOC\.design", js_function_body(js, "readConfig") or ""):
        dynamic = set(k for k, _lab, _memo in core.DESIGN_DOC)
    gone = [k for k in need if k not in (literal | dynamic)]

    # 카드에서 오는 두 키는 GLUE 가 cfg 에 실제로 대입하는지 봅니다.
    # 문자열이 있는지만 보면 cfg["batch_chain"] 을 읽기만 해도 통과해 버립니다.
    miss = [k for k in GLUE_SUPPLIED_KEYS
            if k in core.DESIGN_KEYS and not re.search(
                r"cfg\[\s*[\"']" + re.escape(k) + r"[\"']\s*\]\s*=[^=]", glue or "")]

    report.ok("B", "readConfig 가 DESIGN_KEYS 를 덮는가", not gone and not miss,
              "readConfig %d 키 (근거 %s) + GLUE %d 키"
              % (len(need), "+".join(source), len(GLUE_SUPPLIED_KEYS)),
              "readConfig 가 안 보냄: %s / GLUE 가 안 넣음: %s"
              % (", ".join(gone) or "-", ", ".join(miss) or "-"))


# 이슈 2 에서 CONST / RULES 로 옮긴 옛 모듈 상수. 이름이 남아 있으면 이동이 덜 끝난 것.
MOVED_AWAY = ("_AL_MATCH", "_AL_MIS", "_AL_GAP", "_AL_FLANK",
              "_REPEAT_MAX_PERIOD", "_EXO_TRIM", "_OVERLONG_ZONE", "_FR4_MOTIF")


def sheets_for_check(core, reg_out):
    """시트 구조 검사용. 실행 결과가 있으면 그걸 쓰고, 없으면 빈 입력으로 만든다."""
    if reg_out is not None and reg_out.get("sheets"):
        return reg_out["sheets"]
    cfg = core.build_config(None, [])
    return core.build_sheets([], [], cfg, core.compose([], cfg), {})


def _str_options(node):
    """상수 문자열, 또는 삼항식의 양쪽 상수 문자열."""
    if isinstance(node, ast.Constant) and isinstance(node.value, str):
        return {node.value}
    if isinstance(node, ast.IfExp):
        return _str_options(node.body) | _str_options(node.orelse)
    return set()


def flags_appended_in(fn):
    """함수 안에서 flags 리스트에 실제로 들어가는 문자열 집합."""
    alias = {}
    for node in ast.walk(fn):
        if (isinstance(node, ast.Assign) and len(node.targets) == 1
                and isinstance(node.targets[0], ast.Name)):
            vals = _str_options(node.value)
            if vals:
                alias.setdefault(node.targets[0].id, set()).update(vals)
    out = set()
    for node in ast.walk(fn):
        if (isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute)
                and node.func.attr == "append"
                and isinstance(node.func.value, ast.Name)
                and node.func.value.id == "flags" and node.args):
            vals = _str_options(node.args[0])
            if not vals and isinstance(node.args[0], ast.Name):
                vals = alias.get(node.args[0].id, set())
            out |= vals
    return out


def check_flag_sev_keys(report, core, core_tree):
    """FLAG_SEV 의 키와 qc_one 이 실제로 붙이는 플래그가 양방향으로 맞는가.

    FLAG_SEV 에만 있으면 아무 데도 안 붙는 유령 플래그이고, qc_one 에만 있으면
    심각도 0 으로 취급되어 절대 verdict 가 되지 못합니다. 둘 다 조용한 결함입니다.
    """
    fn = find_function(core_tree, "qc_one")
    if fn is None:
        report.add("B", "FLAG_SEV 대 qc_one 의 실제 플래그", FAIL,
                   "core.py 에서 qc_one 을 찾지 못했습니다")
        return
    used = flags_appended_in(fn)
    sev = set(core.RULES["FLAG_SEV"])
    ghost = sorted(sev - used)
    orphan = sorted(used - sev)
    report.ok("B", "FLAG_SEV 대 qc_one 의 실제 플래그", not ghost and not orphan,
              "%d 개 양방향 일치" % len(sev),
              "FLAG_SEV 에만 있음(유령): %s / qc_one 에만 있음(심각도 0): %s"
              % (", ".join(ghost) or "-", ", ".join(orphan) or "-"))


def check_analysis_mode(report, core):
    """analysis_mode 가 설계 키로 등록되고, MODE_OPTS 값이 그대로 통하는가."""
    bad = []
    default = core.DESIGN_DEFAULT_MAP.get("analysis_mode")
    if "analysis_mode" not in core.DESIGN_KEYS:
        bad.append("DESIGN_DEFAULTS 에 analysis_mode 없음")
    if default not in core.MODE_OPTS:
        bad.append("기본값 %r 이 MODE_OPTS 에 없음" % default)
    for m in core.MODE_OPTS:
        got = core.build_config({"analysis_mode": m}, [])["analysis_mode"]
        if got != m:
            bad.append("%r 를 넘겼는데 %r 로 저장됨" % (m, got))
    if "analysis_mode" not in core.JUDGMENT_DESIGN_KEYS:
        bad.append("JUDGMENT_DESIGN_KEYS 에 analysis_mode 없음")
    # index.html 은 modeOpts 의 순서로 탭 이름을 붙이므로 순서가 계약입니다.
    # 앞 두 값의 순서가 바뀌면 기존 design_hash 도 깨집니다.
    want = [core.MODE_ASSIGNED, core.MODE_LIBRARY, core.MODE_NEGCTRL]
    if list(core.MODE_OPTS) != want:
        bad.append("MODE_OPTS 가 %s 가 아님" % want)
    report.ok("B", "analysis_mode 모드 값", not bad,
              "기본값 %s · MODE_OPTS %s · design_hash 대상"
              % (default, "/".join(core.MODE_OPTS)),
              " / ".join(bad))


def check_ambiguity_runtime(report, core):
    """모호성 항목이 프라이머 이름을 하드코딩하고 있지 않은가.

    프라이머가 없을 때의 용어설명에 프라이머 이름이 남아 있으면 런타임 계산이
    아니라 코드에 박아둔 것입니다. 프라이머를 넣었을 때만 이름이 나와야 합니다.
    """
    # 이번 이슈의 대상은 "모호성 클러스터" 항목입니다. 같은 구분의 다른 항목
    # (Δ · 모호성 표기 · CDR3 경계 · Over 프라이머)은 프라이머 이름을 예시로
    # 인용하고 있고 그건 별개 사안이라 여기서 보지 않습니다.
    TARGET = "모호성 클러스터"
    empty = [d for _c, t, d in core.glossary() if t == TARGET]
    names = set()
    src = os.path.join(ROOT, TESTDATA, "scFv_primers.fa")
    pairs = []
    if os.path.isfile(src):
        with io.open(src, encoding="utf-8", errors="ignore") as fh:
            primers, _t, _w = core.parse_primer_fasta(fh.read())
        names = set(p["name"] for p in primers)
        cfg = core.build_config(None, [])
        pairs = core.primer_ambiguity(primers, cfg)
    leaked = sorted(n for n in names if any(n in d for d in empty))
    bad = []
    if leaked:
        bad.append("프라이머 없이 만든 용어설명에 이름이 남아 있음: " +
                   ", ".join(leaked[:6]))
    if names and not pairs:
        bad.append("프라이머를 넣어도 모호성 쌍이 하나도 계산되지 않았습니다")
    if names:
        filled = [d for _c, t, d in core.glossary(primers, cfg) if t == TARGET]
        if filled and not any(ch.isdigit() for ch in filled[0]):
            bad.append("계산된 용어설명에 수치가 없습니다")
    report.ok("B", "모호성 항목이 런타임 계산인가", not bad,
              "프라이머 없이는 이름 0 건 · 넣으면 %d 쌍 계산" % len(pairs),
              " / ".join(bad))


def check_negctrl_vocab(report, core, html):
    """대조군 verdict 가 전부 용어설명과 index.html badge 에서 처리되는가."""
    verdicts = list(core.NEGCTRL_VERDICTS)
    terms = set(t for cat, t, _d in core.glossary() if cat == "대조군 판정")
    no_doc = [v for v in verdicts if v not in terms]
    no_level = [v for v in verdicts if v not in core.NEGCTRL_LEVEL]
    # index.html 은 등급 -> CSS 클래스만 들고 있습니다. 그 표에 빠진 등급이 있으면
    # 해당 verdict 가 기본색으로 떨어집니다.
    m = re.search(r"const\s+NEG_CLS\s*=\s*\{([^}]*)\}", html)
    cls_keys = set(re.findall(r"([A-Za-z_]\w*)\s*:", m.group(1))) if m else set()
    levels = sorted(set(core.NEGCTRL_LEVEL[v] for v in verdicts if v in core.NEGCTRL_LEVEL))
    no_cls = ([] if m else ["NEG_CLS 를 찾지 못함"]) + \
             [lv for lv in levels if lv not in cls_keys]
    bad = (["용어설명 없음: " + ", ".join(no_doc)] if no_doc else []) + \
          (["등급 없음: " + ", ".join(no_level)] if no_level else []) + \
          (["badge 등급 미처리: " + ", ".join(no_cls)] if no_cls else [])
    report.ok("B", "대조군 verdict 처리", not bad,
              "%d 종 · 용어설명 %d · 등급 %s 전부 badge 처리"
              % (len(verdicts), len(verdicts), "/".join(levels)),
              " / ".join(bad))


def check_judgment_design_keys(report, core):
    """design_hash 대상 키가 실제로 존재하는 설계 키인지. 유령 키를 막는다."""
    keys = list(core.JUDGMENT_DESIGN_KEYS)
    ghost = [k for k in keys if k not in core.DESIGN_KEYS]
    report.ok("B", "JUDGMENT_DESIGN_KEYS 가 DESIGN_DEFAULTS 에 존재", not ghost,
              "%d 키 전부 존재 (%s)" % (len(keys), ", ".join(keys)),
              "DESIGN_DEFAULTS 에 없는 키: " + ", ".join(ghost))


def check_rules_doc(report, core):
    a = set(core.RULES)
    b = set(k for k, _d in core.RULES_DOC)
    report.ok("B", "RULES 대 RULES_DOC 키", a == b,
              "%d 키 일치" % len(a),
              "RULES 만: %s / RULES_DOC 만: %s"
              % (sorted(a - b) or "-", sorted(b - a) or "-"))


def check_moved_constants(report, core):
    left = [n for n in MOVED_AWAY if hasattr(core, n)]
    src = read_text("core.py")
    textual = [n for n in MOVED_AWAY if n in src]
    bad = sorted(set(left) | set(textual))
    report.ok("B", "옛 모듈 상수 잔존", not bad,
              "%d 개 전부 CONST/RULES 로 이동 완료" % len(MOVED_AWAY),
              "core.py 에 남아 있음: " + ", ".join(bad))


def check_rules_exposed(report, core, glue, reg_out):
    """RULES 가 화면(js_docs)과 05_실행설정 시트에 실제로 노출되는지."""
    keys = [k for k, _d in core.RULES_DOC]
    # 시트 조립은 guard 안에서 해야 core.py 가 깨졌을 때도 표를 끝까지 낸다.
    sheets = sheets_for_check(core, reg_out)

    doc_ok, doc_note = False, "GLUE 의 js_docs 를 읽지 못했습니다"
    if glue:
        fn = find_function(ast.parse(glue), "js_docs")
        if fn is not None:
            d = _returned_dict(fn.body[-1].value) if isinstance(fn.body[-1], ast.Return) else None
            got = set(_dict_literal_keys(d)) if d is not None else set()
            hit = sorted(k for k in got if "rule" in k.lower())
            doc_ok = bool(hit)
            doc_note = ("js_docs 키 " + ", ".join(hit)) if hit else \
                       "js_docs 반환에 RULES 관련 키가 없습니다"
    report.ok("B", "js_docs 가 RULES 를 내보내는가", doc_ok, doc_note, doc_note)

    rows = []
    for sh in sheets:
        if str(sh.get("title", "")).startswith("05"):
            rows = sh.get("rows") or []
            break
    listed = [r[1] for r in rows if len(r) > 1 and "알고리즘" in str(r[0])]
    gone = [k for k in keys if k not in listed]
    report.ok("B", "05_실행설정 에 RULES 행", rows and not gone,
              "'고정 상수 (알고리즘)' 행 %d 건 · RULES %d 키 전부 기록"
              % (len(listed), len(keys)),
              ("05_실행설정 시트를 찾지 못했습니다" if not rows
               else "시트에 없는 RULES 키: " + ", ".join(gone)))


def check_cfg_doc(report, core):
    a = set(k for k, _v in core.CFG_DEFAULTS)
    b = set(core.CFG_DOC)
    report.ok("B", "CFG_DEFAULTS 대 CFG_DOC 키", a == b,
              "%d 키 일치" % len(a),
              "DEFAULTS 만: %s / DOC 만: %s"
              % (sorted(a - b) or "-", sorted(b - a) or "-"))


def check_design_doc(report, core):
    a = [k for k, _v in core.DESIGN_DEFAULTS]
    b = set(k for k, _lab, _memo in core.DESIGN_DOC)
    gap = [k for k in a if k not in b]
    new = [k for k in gap if k not in KNOWN_DESIGN_DOC_GAP]
    extra = sorted(b - set(a))
    known = not new and not extra
    if new:
        detail = "DESIGN_DOC 에 없는 새 키 %d 건: %s" % (len(new), ", ".join(new))
    elif extra:
        detail = "DESIGN_DEFAULTS 에 없는데 DOC 에만 있음: " + ", ".join(extra)
    elif gap:
        detail = "DEFAULTS %d · DOC %d · 알려진 누락 %d 건 (%s)" % (
            len(a), len(b), len(gap), ", ".join(gap))
    else:
        detail = "%d 키 일치" % len(a)
    report.add("B", "DESIGN_DEFAULTS 대 DESIGN_DOC 키", PASS if known else FAIL, detail)


# =============================================================================
#  [C] 누출
# =============================================================================
_SEQ_RE = re.compile(r"(?<![A-Za-z0-9_])[ACGT]{6,}(?![A-Za-z0-9_])")


def check_no_sequence(report, html):
    stripped = html.replace(MARKSEQ_REGEX, "")
    hits = []
    for m in _SEQ_RE.finditer(stripped):
        line = stripped.count("\n", 0, m.start()) + 1
        hits.append("%d행 %s" % (line, m.group(0)))
    report.ok("C", "index.html 6nt 이상 ACGT 리터럴", not hits,
              "0 건 (markSeq 정규식은 문자 클래스이므로 제외)",
              "발견: " + " / ".join(hits[:4]))


def _html_variants(text):
    """index.html 이 · 나 — 를 엔티티로 써도 잡히도록 표기 변형을 만든다."""
    ent = text.replace("·", "&middot;").replace("—", "&mdash;")
    return {text, ent}


def check_no_label_leak(report, html, core):
    """설정 라벨은 전부 core.py 에서 읽어야 한다. index.html 에 리터럴로 있으면 실패."""
    labels = [("DESIGN_DOC", lab) for _k, lab, _m in core.DESIGN_DOC]
    labels += [("CFG_DOC", core.CFG_DOC[k][0]) for k in core.THRESH_KEYS]
    hits = []
    for src, lab in labels:
        if lab and any(v in html for v in _html_variants(lab)):
            hits.append("%s '%s'" % (src, lab))
    report.ok("C", "설정 라벨이 index.html 에 등장", not hits,
              "DESIGN_DOC %d + CFG_DOC %d 라벨 전부 미등장"
              % (len(core.DESIGN_DOC), len(core.THRESH_KEYS)),
              "index.html 에 하드코딩됨: " + ", ".join(hits))


def check_no_mode_leak(report, html, core):
    """분석 모드 값은 core 에서 읽어야 합니다. index.html 에 리터럴로 있으면 실패.

    서열·라벨과 달리 모드 값은 식별자라서 analyze 반환 키(M.negctrl)나 core 함수
    이름(negctrl_summary)처럼 하드코딩이 아닌 곳에도 같은 글자가 나옵니다. 그래서
    원문 전체가 아니라 따옴표 안의 문자열 리터럴만 봅니다 — 모드 값을 코드에
    박아 넣으려면 반드시 문자열 리터럴이어야 하기 때문입니다.
    """
    lits = js_strings(html)
    hits = [m for m in core.MODE_OPTS if m and any(m in x for x in lits)]
    report.ok("C", "모드 값이 index.html 에 등장", not hits,
              "MODE_OPTS %d 개 전부 문자열 리터럴에 없음 (js_docs 의 modeOpts 로 전달)"
              % len(core.MODE_OPTS),
              "index.html 에 하드코딩됨: " + ", ".join(hits))


def check_no_const_leak(report, html, core):
    hits = []
    for k, v in core.CONST.items():
        if isinstance(v, str) and v and v in html:
            hits.append(k)
    report.ok("C", "CONST 서열이 index.html 에 등장", not hits,
              "문자열 상수 %d 개 전부 미등장"
              % sum(1 for v in core.CONST.values() if isinstance(v, str)),
              "index.html 에 등장: " + ", ".join(hits))


# =============================================================================
#  [D] 회귀
# =============================================================================
def testdata_inputs():
    d = os.path.join(ROOT, TESTDATA)
    if not os.path.isdir(d):
        return None, None
    ab1 = sorted(n for n in os.listdir(d) if n.lower().endswith(".ab1"))
    fa = [n for n in os.listdir(d) if n.lower().endswith((".fa", ".fasta", ".fas"))]
    if not ab1 or not fa:
        return None, None
    files = []
    for n in ab1:
        with open(os.path.join(d, n), "rb") as fh:
            files.append((n, fh.read()))
    pick = "scFv_primers.fa" if "scFv_primers.fa" in fa else sorted(fa)[0]
    with io.open(os.path.join(d, pick), encoding="utf-8", errors="ignore") as fh:
        return files, fh.read()


def norm_cell(v):
    """xlsx_writer 가 셀에 쓰는 값과 openpyxl 이 읽어온 값을 같은 자리로 옮긴다."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v.is_integer() and abs(v) < 1e15:
            return int(v)
        return round(v, 10)
    return str(v)


D_NAMES = ["param_hash", "design_hash", "클론별 판정·길이", "클론별 혼합(%)", "CDR3-H3 중앙값",
           "조성 VL 순서", "조성 JH 순서", "fasta 통과 클론 수",
           "xlsx 바이트 재현성", "xlsx openpyxl 재검증"]


def run_analysis(core):
    """[B] 와 [D] 가 함께 쓰는 analyze 실행. (out, 사유, 상태) 를 반환한다.

    [B] 는 이 결과의 실제 키를 대조에 쓴다. analyze 에는 오류 경로와 성공 경로의
    return 이 따로 있어, AST 로 반환 키를 읽으면 엉뚱한 쪽을 볼 수 있기 때문이다.
    """
    files, ptext = testdata_inputs()
    if files is None:
        return None, "testdata/ 에 .ab1 또는 프라이머 FASTA 가 없습니다", SKIP
    meta = {"runtime": "verify.py " + VERIFY_VERSION,
            "timestamp": "0000-00-00 00:00:00", "primer_file": "scFv_primers.fa"}
    meta.update(REG_META)
    try:
        out = core.analyze(files, ptext, dict(REG_OVERRIDES), meta)
    except Exception as e:
        return None, "analyze 예외 %s: %s" % (type(e).__name__, e), FAIL
    if not out.get("ok"):
        return None, "analyze 중단: " + "; ".join(out.get("errors", [])), FAIL
    return out, "", PASS


def check_regression(report, out, note, status):
    if out is None:
        for n in D_NAMES:
            report.add("D", n, status, note)
        return

    import xlsx_writer

    for name in ("param_hash", "design_hash"):
        got = out["config"].get(name)
        report.ok("D", name, got == EXPECT[name], EXPECT[name],
                  "기대 %s 실제 %s" % (EXPECT[name], got))

    qc = dict((q["id"], q) for q in out["qc"])
    bad = []
    for e in EXPECT["clones"]:
        q = qc.get(e["id"])
        if q is None:
            bad.append("%s (%s) 없음" % (e["tag"], e["id"]))
            continue
        if q["verdict"] != e["verdict"]:
            bad.append("%s verdict %s(기대 %s)" % (e["tag"], q["verdict"], e["verdict"]))
        if q["insert_bp"] != e["insert"]:
            bad.append("%s insert %s(기대 %d)" % (e["tag"], q["insert_bp"], e["insert"]))
        for f in ("d1", "d2"):
            if f in e and q[f] != e[f]:
                bad.append("%s %s %s(기대 %d)" % (e["tag"], f, q[f], e[f]))
    report.ok("D", "클론별 판정·길이", not bad,
              "%d 클론 · verdict / insert / d1 / d2 일치" % len(EXPECT["clones"]),
              " / ".join(bad))

    badm = []
    for e in EXPECT["clones"]:
        q = qc.get(e["id"]) or {}
        got = None if q.get("mix_pct") is None else round(q["mix_pct"], 1)
        if got != e["mix"]:
            badm.append("%s %s(기대 %s)" % (e["tag"], got, e["mix"]))
    report.ok("D", "클론별 혼합(%)", not badm,
              " / ".join("%s %.1f" % (e["tag"], e["mix"]) for e in EXPECT["clones"]),
              " / ".join(badm))

    comp = out["composition"]
    report.ok("D", "CDR3-H3 중앙값", comp["cdr3_median"] == EXPECT["cdr3_median"],
              str(EXPECT["cdr3_median"]),
              "기대 %s 실제 %s" % (EXPECT["cdr3_median"], comp["cdr3_median"]))
    for label, key in (("조성 VL 순서", "vl"), ("조성 JH 순서", "jh")):
        got = [tuple(x) for x in comp[key]]
        want = [tuple(x) for x in EXPECT[key]]
        report.ok("D", label, got == want, str(want), "기대 %s 실제 %s" % (want, got))

    n_fa = out["fasta"].count(">")
    report.ok("D", "fasta 통과 클론 수", n_fa == EXPECT["fasta_n"],
              str(EXPECT["fasta_n"]),
              "기대 %d 실제 %d" % (EXPECT["fasta_n"], n_fa))

    sheets = out["sheets"]
    b1 = xlsx_writer.write_xlsx(sheets)
    b2 = xlsx_writer.write_xlsx(sheets)
    report.ok("D", "xlsx 바이트 재현성", b1 == b2,
              "두 번 호출 결과 동일 · %d bytes" % len(b1),
              "두 번 호출 결과가 다릅니다 (%d vs %d bytes)" % (len(b1), len(b2)))

    check_xlsx_roundtrip(report, sheets, b1)


def check_xlsx_roundtrip(report, sheets, data):
    try:
        import openpyxl
    except ImportError:
        report.add("D", "xlsx openpyxl 재검증", SKIP, "openpyxl 없음")
        return
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    bad, ncell = [], 0
    if len(wb.worksheets) != len(sheets):
        report.add("D", "xlsx openpyxl 재검증", FAIL,
                   "시트 수 기대 %d 실제 %d" % (len(sheets), len(wb.worksheets)))
        return
    for si, sheet in enumerate(sheets):
        ws = wb.worksheets[si]
        headers = list(sheet.get("headers") or [])
        rows = list(sheet.get("rows") or [])
        note = sheet.get("note")
        ncol = max([len(headers)] + [len(r) for r in rows]) if (headers or rows) else 1
        top = 3 if note else 1
        want = []
        if note:
            want.append((1, [note] + [None] * (ncol - 1)))
        if headers:
            want.append((top, list(headers) + [None] * (ncol - len(headers))))
        for i, row in enumerate(rows):
            want.append((top + 1 + i, list(row) + [None] * (ncol - len(row))))
        for rn, vals in want:
            for j, v in enumerate(vals, start=1):
                ncell += 1
                got = norm_cell(ws.cell(row=rn, column=j).value)
                exp = norm_cell(v)
                if got != exp:
                    bad.append("%s!%s%d 기대 %r 실제 %r"
                               % (sheet["title"], openpyxl.utils.get_column_letter(j),
                                  rn, exp, got))
    report.ok("D", "xlsx openpyxl 재검증", not bad,
              "%d 시트 · %d 셀 전부 일치" % (len(sheets), ncell),
              " / ".join(bad[:3]) + (" 외 %d" % (len(bad) - 3) if len(bad) > 3 else ""))


# =============================================================================
#  [E] 단위 — 실측 testdata 가 밟지 않는 랜드마크 분기
# =============================================================================
# testdata 4 클론의 랜드마크 gap 은 {0, 1} 뿐이라 GAP# 분기가 한 번도 실행되지
# 않습니다. 합성 서열로 check_landmark 를 직접 불러 그 구간을 덮습니다.
UNIT_PRE = "T" * 20                 # 랜드마크와 무관한 5' 패딩
UNIT_SUF = "A" * 20                 # 랜드마크와 무관한 3' 패딩
UNIT_ALIEN = ("ATTC" * 12)[:45]     # 랜드마크 자리에 넣을 무관한 45 nt
UNIT_SUB2 = (6, 22)                 # 치환 위치 — 허용치(lm_max_sub) 이내
UNIT_SUB3 = (6, 22, 38)             # 치환 위치 — 허용치 초과
UNIT_DEL_AT = 12                    # 결실 시작 위치 (동종중합 구간 밖)
# 5' G-run 경계. 여기서 2 nt 를 지우면 갭 1 개와 치환 2 개가 박빙이라
# AL_GAP 이 -3 이 되면 S2G1/WARN 으로 내려앉는다. AL_GAP 회귀 감지용.
# 같은 위치라도 n=1 / n=3 은 격차가 커서 갈리지 않으므로 n=2 만 쓴다.
UNIT_DEL_TIGHT = 3
UNIT_INS_AT = 12                    # 삽입 위치
# 삽입 염기. QC2 의 최장 C 연속은 2 라 6~12 nt C 런은 랜드마크와 무관하다.
# (45 nt 로 환산했을 때 최소 불일치 37 > 허용 치환 2 — 아래 무관성 검사에서 확인)
UNIT_INS_UNIT = "C"
_TRANSITION = {"A": "G", "G": "A", "C": "T", "T": "C"}

#        태그    설명                      변형종류  인자               기대 status  기대 level
UNIT_CASES = [
    ("E1", "변형 없음",             "sub",   (),                    "OK",     "OK"),
    ("E2", "치환 2 개 (허용 이내)", "sub",   UNIT_SUB2,             "OK",     "OK"),
    ("E3", "치환 3 개 (허용 초과)", "sub",   UNIT_SUB3,             "S3G0",   "WARN"),
    ("E4", "1 nt 결실",             "del",   (UNIT_DEL_AT, 1),      "S0G1",   "WARN"),
    ("E5", "2 nt 결실",             "del",   (UNIT_DEL_AT, 2),      "GAP2",   "FAIL"),
    ("E6", "3 nt 결실",             "del",   (UNIT_DEL_AT, 3),      "GAP3",   "FAIL"),
    ("E7", "무관한 서열",           "alien", None,                  "ABSENT", "FAIL"),
    ("E8", "covered=False",         "na",    None,                  "NA",     "NA"),
    ("E9", "결실 위치 3 · 2 nt 결실", "del", (UNIT_DEL_TIGHT, 2),   "GAP2",   "FAIL"),
    # AL_FLANK 가 12 미만이면 ABSENT 로 떨어진다. AL_FLANK 회귀 감지용.
    # 기대값은 RULES 값에 연동하지 않고 GAP6 으로 고정한다. 연동하면 상수가
    # 바뀔 때 기대값도 따라 바뀌어 회귀를 못 잡는다.
    ("E10", "6 nt 삽입",            "ins",   (UNIT_INS_AT, 6),      "GAP6",   "FAIL"),
    # flank 12 로도 담기지 않는 크기. 창을 넘어선 삽입의 판정을 고정한다.
    ("E11", "12 nt 삽입",           "ins",   (UNIT_INS_AT, 12),     "ABSENT", "FAIL"),
]


def mutate_landmark(lm, kind, arg):
    if kind == "sub":
        b = list(lm)
        for p in arg:
            b[p] = _TRANSITION[b[p]]
        return "".join(b)
    if kind == "del":
        pos, n = arg
        return lm[:pos] + lm[pos + n:]
    if kind == "ins":
        pos, n = arg
        return lm[:pos] + (UNIT_INS_UNIT * n)[:n] + lm[pos:]
    if kind == "alien":
        return UNIT_ALIEN
    return lm


def check_landmark_units(report, core):
    cfg = core.build_config(None, [])
    lm = core.CONST["QC2"]

    # 패딩·대체서열·삽입서열이 랜드마크와 우연히 맞지 않는지 먼저 확인한다.
    scores, bad = [], []
    for name, mid in (("대체", UNIT_ALIEN),
                      ("삽입", (UNIT_INS_UNIT * len(lm))[:len(lm)])):
        su, _pu = core.ungapped_scan(lm, UNIT_PRE + mid + UNIT_SUF)
        scores.append("%s %d" % (name, su))
        if su <= cfg["lm_max_sub"]:
            bad.append("%s 서열이 랜드마크와 맞습니다 (불일치 %d ≤ 허용 %d)"
                       % (name, su, cfg["lm_max_sub"]))
    report.ok("E", "패딩·대체·삽입 서열 무관성", not bad,
              "최소 불일치 " + " / ".join(scores) + " > 허용 치환 %d" % cfg["lm_max_sub"],
              " / ".join(bad))

    for tag, label, kind, arg, want_st, want_lv in UNIT_CASES:
        seq = UNIT_PRE + mutate_landmark(lm, kind, arg) + UNIT_SUF
        r = core.check_landmark(lm, seq, [], kind != "na", cfg)
        good = (r["status"] == want_st and r["level"] == want_lv)
        detail = "status %s · level %s" % (want_st, want_lv)
        if tag == "E1":
            good = good and r["pos"] == len(UNIT_PRE)
            detail += " · 위치 %d" % r["pos"]
        report.ok("E", tag + " " + label, good, detail,
                  "기대 %s/%s · 실제 %s/%s (sub %d, gap %d, pos %d)"
                  % (want_st, want_lv, r["status"], r["level"],
                     r["sub"], r["gap"], r["pos"]))


# --- 스터퍼 길이 판정 (PARENTAL / PARENTAL?) 합성 시험 -------------------------
# 실측 testdata 에는 인서트 386 bp 인 클론이 없어 [D] 로는 이 분기가 한 번도
# 실행되지 않습니다. .ab1 없이 qc_one 을 돌리려면 raw_seq / raw_qual / trace /
# ploc 을 갖춘 dict 면 충분합니다. raw_qual=[] 이면 트리밍을 건너뛰고,
# trace={} 이면 mix_pct 가 None 이라 MIXED 가 붙지 않습니다. 둘 다 이 시험의
# 판정 대상이 아니므로 결과에 영향이 없습니다.
UNIT_FILL = "ATTC"          # NotI / AscI / 링커와 무관한 채움 서열


def synth_stuffer_read(core, with_stuffer):
    """인서트(NotI 첫 염기 ~ AscI 끝)가 정확히 STUFFER_INSERT_BP 인 합성 read."""
    C = core.CONST
    q1off = C["QC1"].find(C["NotI"])
    span = len(C["QC1"]) - q1off + len(C["AscI"])     # NotI 시작 ~ QC1 끝, 그리고 AscI
    mid_len = C["STUFFER_INSERT_BP"] - span
    filler = UNIT_FILL * (mid_len // len(UNIT_FILL) + 2)
    mid = (C["STUFFER"] + filler)[:mid_len] if with_stuffer else filler[:mid_len]
    seq = C["PELB_ATG"] + "GGG" + C["QC1"] + mid + C["QC3"] + C["QC4"] + "A" * 20
    return {"id": "syn", "clone": "syn", "batch": "", "date": "", "primer": "",
            "filename": "synthetic.ab1", "raw_len": len(seq),
            "raw_seq": seq, "raw_qual": [], "trace": {}, "ploc": []}


def check_stuffer_units(report, core):
    cfg = core.build_config(None, [])
    want_bp = core.CONST["STUFFER_INSERT_BP"]
    cases = [("E12", "insert %d · 스터퍼 서열 있음" % want_bp, True, "PARENTAL", "PARENTAL?"),
             ("E13", "insert %d · 스터퍼 서열 없음" % want_bp, False, "PARENTAL?", "PARENTAL")]
    for tag, label, with_stuffer, want, unwanted in cases:
        r = core.qc_one(synth_stuffer_read(core, with_stuffer), cfg)
        good = (r["insert_bp"] == want_bp
                and want in r["flags"] and unwanted not in r["flags"])
        report.ok("E", tag + " " + label, good,
                  "insert %s · %s 포함 · %s 미포함 · verdict %s"
                  % (r["insert_bp"], want, unwanted, r["verdict"]),
                  "insert %s (기대 %d) · flags %s (%s 포함 / %s 미포함 이어야 함)"
                  % (r["insert_bp"], want_bp, r["flags"], want, unwanted))


# --- 대조군 판정 분기 합성 시험 ------------------------------------------------
# 실측 대조군으로는 두 분기 정도만 실행됩니다. 나머지는 합성 read 로 덮습니다.
# (변형종류, 인자, 기대 verdict)
NEG_CASES = [
    ("N1", "스터퍼 서열 검출",           "stuffer", "PARENTAL"),
    ("N2", "인서트가 스터퍼 길이",        "len386",  "PARENTAL?"),
    ("N3", "링커 O · 범위 내 · 프레임 정상", "scfv",  "CONTAMINATED"),
    ("N4", "인서트 있으나 링커 없음",     "nolink",  "CARRYOVER"),
    ("N5", "AscI 미검출",                "noascI",  "EMPTY_VECTOR"),
    ("N6", "NotI 2 회",                  "concat",  "CONCATEMER"),
    ("N7", "트레이스 혼합",              "mixed",   "MIXED"),
    ("N8", "어느 분류에도 안 맞음",       "check",   "CHECK"),
    ("N9", "링커 O · 범위 내 · 프레임 이상", "offframe", "CONTAMINATED?"),
    ("N10", "링커 O · 인서트 하한 미만",   "partial", "PARTIAL_INSERT"),
]


def synth_neg_read(core, cfg, kind):
    """대조군 분기용 합성 read. 이슈 6 의 합성 read 방식을 그대로 씁니다.

    입력 길이는 cfg 에서 계산합니다(기대값이 아니라 입력이므로 연동해도 됩니다).
    기대 verdict 는 NEG_CASES 에 문자열로 고정되어 있습니다.
    """
    C = core.CONST
    q1off = C["QC1"].find(C["NotI"])
    span = len(C["QC1"]) - q1off + len(C["AscI"])
    fill = UNIT_FILL * 500

    def build(mid, tail=True, extra=""):
        seq = (C["PELB_ATG"] + "GGG" + C["QC1"] + mid +
               (C["QC3"] + C["QC4"] if tail else fill[:60]) + extra + "A" * 20)
        return {"id": "syn", "clone": "syn", "batch": "", "date": "", "primer": "",
                "filename": "syn.ab1", "raw_len": len(seq), "raw_seq": seq,
                "raw_qual": [], "trace": {}, "ploc": []}

    def with_linker(want_ins):
        """링커를 품고 인서트 길이가 want_ins 가 되는 mid."""
        n = want_ins - span - len(C["QC2"])
        half = n // 2
        return build(fill[:half] + C["QC2"] + fill[:n - half])

    if kind == "stuffer":                      # 스터퍼 + 스터퍼 길이
        return build((C["STUFFER"] + fill)[:C["STUFFER_INSERT_BP"] - span])
    if kind == "len386":                       # 길이만 스터퍼와 같고 서열 없음
        return build(fill[:C["STUFFER_INSERT_BP"] - span])
    if kind in ("scfv", "offframe"):
        want_frame = (kind == "scfv")
        for ins in range(cfg["insert_min"], cfg["insert_max"] + 1):
            if (ins % 3 == C["FRAME_MOD"]) == want_frame:
                return with_linker(ins)
    if kind == "partial":                      # 링커 O · 하한 미만
        return with_linker(cfg["insert_min"] - 200)
    if kind == "nolink":                       # 인서트 있고 링커 없음
        return build(fill[:200])
    if kind == "noascI":                       # AscI 없음
        return build(fill[:200], tail=False)
    if kind == "concat":                       # NotI 2 회
        return build(fill[:200], extra=C["NotI"])
    if kind == "check":                        # 링커 있고 인서트가 상한 초과
        return with_linker(cfg["insert_max"] + 120)
    return build(fill[:200])


def check_negctrl_units(report, core):
    cfg = core.build_config({"analysis_mode": core.MODE_NEGCTRL}, [])
    for tag, label, kind, want in NEG_CASES:
        if kind == "mixed":
            # 트레이스가 없으면 mix_pct 가 None 이라 이 분기에 닿지 않습니다.
            r = dict(core.qc_one(synth_neg_read(core, cfg, "nolink"), cfg))
            r["mix_pct"] = cfg["mix_pct"] + 1.0
        else:
            r = core.qc_one(synth_neg_read(core, cfg, kind), cfg)
        got, why = core.negctrl_verdict(r, cfg)
        # 모든 reason 에 인서트 길이와 프레임이 들어가야 합니다.
        has_facts = "인서트 " in why and "프레임" in why
        report.ok("E", "%s %s" % (tag, label), got == want and has_facts,
                  "%s · %s" % (want, why[:30]),
                  "기대 %s 실제 %s%s (%s)"
                  % (want, got, "" if has_facts else " · reason 에 인서트/프레임 없음",
                     why))


# --- index.html 의 badge() 가 아래 표시값을 모두 처리하는가 --------------------
BADGE_EXPECT = [("OK", "b-pass"), ("S1G0", "b-warn"), ("GAP2", "b-fail"),
                ("ABSENT", "b-fail"), ("NA", "b-none"), ("PARENTAL?", "b-check")]

_BADGE_RULE_RE = re.compile(
    r"^\s*if\s*\((?P<cond>.+?)\)\s*return\s*'<span class=\"badge (?P<cls>b-[a-z]+)\"")
_BADGE_DEFAULT_RE = re.compile(
    r"^\s*return\s*'<span class=\"badge (?P<cls>b-[a-z]+)\"")


def _js_unquote(text):
    return text.replace('\\"', '"').replace("\\\\", "\\")


def eval_badge_cond(cond, s):
    """badge() 의 조건식을 제한 해석한다. 해석하지 못하면 None."""
    for part in cond.split("||"):
        part = part.strip()
        m = re.fullmatch(r's\s*===\s*"((?:[^"\\]|\\.)*)"', part)
        if m:
            if s == _js_unquote(m.group(1)):
                return True
            continue
        m = re.fullmatch(r"/(.+)/[a-z]*\.test\(s\)", part)
        if m:
            try:
                if re.search(m.group(1), s):
                    return True
            except re.error:
                return None
            continue
        return None
    return False


def badge_class(body, s):
    """badge() 본문을 위에서부터 따라가며 s 가 받을 클래스와 명시/기본 여부."""
    for line in body.splitlines():
        m = _BADGE_RULE_RE.match(line)
        if m:
            got = eval_badge_cond(m.group("cond"), s)
            if got is None:
                return None, "해석실패"
            if got:
                return m.group("cls"), "명시"
            continue
        m = _BADGE_DEFAULT_RE.match(line)
        if m:
            return m.group("cls"), "기본값"
    return None, "return 없음"


BADGE_LABEL = "badge() 표시값 %d 종" % len(BADGE_EXPECT)


def check_ambiguity_units(report, core):
    """primer_ambiguity 단위시험. 비호환 위치 수를 0~3 으로 만들어 등급을 확인합니다."""
    cfg = core.build_config(None, [])
    tol = cfg["primer_max_mismatch"]
    base = "ACGTACGTACGTACGTACGTAC"      # 판별구간 22 nt
    swap = {"A": "C", "C": "A", "G": "T", "T": "G"}

    def mk(name, seq, fam):
        return {"name": name, "seq": seq, "core": seq, "core_trim": 0,
                "len": len(seq), "group": "F1_For", "chain": "heavy",
                "family": fam, "families": [fam], "target": "", "fragment": "",
                "dir": "", "tm": ""}

    def mutate(n):
        b = list(base)
        for i in range(n):
            b[i * 3] = swap[b[i * 3]]
        return "".join(b)

    bad = []
    for n in (0, 1, 2, 3, 4, 5):
        pairs = core.primer_ambiguity([mk("p1", base, "F1"), mk("p2", mutate(n), "F1")], cfg)
        want_in = n <= 2 * tol
        if bool(pairs) != want_in:
            bad.append("비호환 %d 이 %s" % (n, "빠짐" if want_in else "포함됨"))
            continue
        if not pairs:
            continue
        p = pairs[0]
        if p["incompatible"] != n:
            bad.append("비호환 %d 인데 %d 로 셈" % (n, p["incompatible"]))
        want_tier = "certain" if n <= tol else "split"
        if p["tie"] != want_tier:
            bad.append("비호환 %d 등급 %s (기대 %s)" % (n, p["tie"], want_tier))

    # 길이가 다르면 짧은 쪽까지만 비교하고 truncated 로 표시
    long_p = mk("p3", base + "GGGGG", "F1")
    pairs = core.primer_ambiguity([mk("p1", base, "F1"), long_p], cfg)
    if not pairs:
        bad.append("길이 다른 쌍이 빠짐")
    else:
        p = pairs[0]
        if p["cmp_len"] != len(base) or not p["truncated"] or p["incompatible"] != 0:
            bad.append("길이 다름: cmp_len %d truncated %s 비호환 %d"
                       % (p["cmp_len"], p["truncated"], p["incompatible"]))

    # family 가 다르면 same_family 가 False
    pairs = core.primer_ambiguity([mk("p1", base, "F1"), mk("p2", base, "F2")], cfg)
    if not pairs or pairs[0]["same_family"] or pairs[0]["families"] != ["F1", "F2"]:
        bad.append("family 다른 쌍을 same_family 로 봄")

    # 부분적으로 겹치는 경우 (예: JH1|JH2 대 JH2). 보고되는 family 는 합집합이라
    # 유일하지 않습니다. 교집합으로 판단하면 이 경우를 같은 family 로 잘못 봅니다.
    p1, p2 = mk("p1", base, "F1"), mk("p2", base, "F2")
    p1["families"], p2["families"] = ["F1", "F2"], ["F2"]
    pairs = core.primer_ambiguity([p1, p2], cfg)
    if not pairs or pairs[0]["same_family"] or pairs[0]["families"] != ["F1", "F2"]:
        bad.append("family 부분 겹침을 same_family 로 봄")

    report.ok("E", "primer_ambiguity 단위", not bad,
              "비호환 0~%d certain · %d~%d split · %d 초과 제외 · 길이 다르면 짧은 쪽까지"
              % (tol, tol + 1, 2 * tol, 2 * tol),
              " / ".join(bad))


def check_badge_states(report, js):
    body = js_function_body(js, "badge")
    if body is None:
        report.add("E", BADGE_LABEL, FAIL, "index.html 에서 badge() 를 찾지 못했습니다")
        return
    bad, shown = [], []
    for s, want in BADGE_EXPECT:
        cls, how = badge_class(body, s)
        shown.append("%s:%s%s" % (s, (cls or how).replace("b-", ""),
                                  "*" if how == "기본값" else ""))
        if cls != want:
            bad.append("%s 기대 %s 실제 %s(%s)" % (s, want, cls, how))
    report.ok("E", BADGE_LABEL, not bad,
              " ".join(shown) + "  (*기본값)", " / ".join(bad))


# =============================================================================
#  [F] GLUE 경로 — 배치별 core.analyze 호출과 병합
# =============================================================================
# [D] 는 js_analyze(단일 호출) 만 덮습니다. 배치 경로는 index.html 의 GLUE 에
# 있어 core 만 불러서는 닿지 않으므로, GLUE 템플릿 리터럴을 꺼내 그대로 실행하고
# testdata 를 두 배치로 나눠 돌립니다. 기대값은 실측해 고정했습니다.
F_NAMES = ["F1 배치 2 개 병합", "F2 배치 label 구분", "F3 배치 지정 불일치",
           "F4 param_hash 불일치 오류", "F5 library 경로", "F6 병합 시트 구조"]

GLUE_EXPECT = {
    "param_hash": "3473927a",
    "design_hash": "8b1eab32",          # VH6 x kappa
    "lib_design_hash": "68210578",      # library + 배치 지정 없음
    "merged_rows": 4,
    "n_good": 2,
    "cdr3_median": 10,
    "fasta_n": 2,
    "ids": [_STEM % "01", _STEM % "02", _STEM % "03", _STEM % "04"],
    "titles": ["01_판정요약", "02_구조QC상세", "03_프라이머판별", "04_배치조성",
               "05_실행설정", "06_용어설명", "07_서열"],
    "sheet01_rows": 4,
    "block04": 3,                       # 배치 블록 2 + 전체 블록 1
}


class _ShimCore(object):
    """F4 전용. 두 번째 analyze 의 param_hash 만 바꿔 배치 간 불일치를 만듭니다.

    공개 API 로는 배치마다 다른 임계값을 줄 수 없어(cfg 가 하나) 이 방법으로만
    js_analyze_batches 의 param_hash 일치 확인 분기에 닿을 수 있습니다.
    """

    def __init__(self, real):
        self._real = real
        self._n = 0

    def __getattr__(self, name):
        return getattr(self._real, name)

    def analyze(self, *a, **kw):
        out = self._real.analyze(*a, **kw)
        self._n += 1
        if self._n == 2 and out.get("ok"):
            out = dict(out)
            out["config"] = dict(out["config"])
            out["config"]["param_hash"] = "0badf00d"
        return out


def _glue_env(glue, root):
    """GLUE 를 실행하고, 파일 입출력만 임시 디렉터리로 돌려놓은 네임스페이스."""
    ns = {}
    keep = list(sys.path)
    try:
        exec(compile(glue, "GLUE(index.html)", "exec"), ns)
    finally:
        sys.path[:] = keep            # GLUE 의 sys.path.insert("/app") 되돌리기
    src = os.path.join(ROOT, TESTDATA)
    names = sorted(n for n in os.listdir(src) if n.lower().endswith(".ab1"))
    fa = sorted(n for n in os.listdir(src) if n.lower().endswith((".fa", ".fasta", ".fas")))
    dirs = {}
    for key, part in (("b1", names[:2]), ("b2", names[2:4]), ("all", names[:4])):
        d = os.path.join(root, key)
        os.makedirs(d)
        for n in part:
            shutil.copyfile(os.path.join(src, n), os.path.join(d, n))
        dirs[key] = d
    with io.open(os.path.join(src, fa[0]), encoding="utf-8", errors="ignore") as fh:
        ptext = fh.read()
    box = {}
    orig = ns["_inputs_from"]
    ns["_primer_text"] = lambda: ptext
    ns["_inputs_from"] = lambda d: orig(dirs["all"] if d == "/input" else d)
    ns["_write_output"] = lambda sheets, headers, rows, fasta: box.update(
        sheets=sheets, headers=headers, rows=rows, fasta=fasta)
    return ns, dirs, box


def _base_cfg(core):
    cfg = dict(core.CFG_DEFAULTS)
    cfg.update(dict(core.DESIGN_DEFAULTS))
    return cfg


def _run_batches(ns, core, specs, cfg, meta):
    return json.loads(ns["js_analyze_batches"](
        json.dumps(specs), json.dumps(cfg), json.dumps(meta)))


def check_glue(report, glue, core):
    if not glue:
        for n in F_NAMES:
            report.add("F", n, SKIP,
                       "index.html 에서 GLUE 템플릿 리터럴을 추출하지 못했습니다")
        return
    src = os.path.join(ROOT, TESTDATA)
    ab1 = sorted(n for n in os.listdir(src)
                 if n.lower().endswith(".ab1")) if os.path.isdir(src) else []
    fa = [n for n in os.listdir(src)
          if n.lower().endswith((".fa", ".fasta", ".fas"))] if os.path.isdir(src) else []
    if len(ab1) < 4 or not fa:
        for n in F_NAMES:
            report.add("F", n, SKIP, "testdata/ 에 .ab1 4 개와 프라이머 FASTA 가 필요합니다")
        return

    root = tempfile.mkdtemp(prefix="scfvqc_glue_")
    try:
        try:
            ns, dirs, box = _glue_env(glue, root)
        except Exception as e:
            for n in F_NAMES:
                report.add("F", n, FAIL,
                           "GLUE 실행 실패 %s: %s" % (type(e).__name__, e))
            return
        cfg = _base_cfg(core)
        meta = {"runtime": "verify.py " + VERIFY_VERSION, "timestamp": "0",
                "primer_file": fa[0], "batch_label": "260819", "batch_date": "260819"}
        same = [{"dir": dirs["b1"], "vh": "VH6", "chain": "kappa"},
                {"dir": dirs["b2"], "vh": "VH6", "chain": "kappa"}]

        out = _run_batches(ns, core, same, cfg, meta)
        if not out.get("ok"):
            for n in F_NAMES[:4] + F_NAMES[5:]:
                report.add("F", n, FAIL, "배치 실행 실패: " +
                           "; ".join(out.get("errors", [])))
            _check_glue_library(report, ns, core, cfg, meta)
            return

        # 뒤이은 실행이 box 를 덮으므로 배치 모드 시트를 여기서 떠 둡니다.
        merged_box = dict(box)
        _check_f1(report, out)
        _check_f2(report, out)
        _check_f3(report, ns, core, dirs, cfg, meta)
        _check_f4(report, ns, core, same, cfg, meta)
        _check_glue_library(report, ns, core, cfg, meta)
        _check_f6(report, merged_box)
    finally:
        shutil.rmtree(root, ignore_errors=True)


def _check_f1(report, out):
    m = out["merged"]
    hs = sorted(set(b["result"]["config"]["param_hash"] for b in out["batches"]))
    ds = sorted(set(b["design_hash"] for b in out["batches"]))
    ids = [r[0] for r in m["summary"]["rows"]]
    bad = []
    if hs != [GLUE_EXPECT["param_hash"]]:
        bad.append("param_hash %s" % hs)
    if ds != [GLUE_EXPECT["design_hash"]]:
        bad.append("design_hash %s" % ds)
    if len(m["summary"]["rows"]) != GLUE_EXPECT["merged_rows"]:
        bad.append("merged rows %d" % len(m["summary"]["rows"]))
    if m["composition"]["n_good"] != GLUE_EXPECT["n_good"]:
        bad.append("n_good %s" % m["composition"]["n_good"])
    if m["composition"]["cdr3_median"] != GLUE_EXPECT["cdr3_median"]:
        bad.append("cdr3_median %s" % m["composition"]["cdr3_median"])
    if m["fasta"].count(">") != GLUE_EXPECT["fasta_n"]:
        bad.append("fasta %d" % m["fasta"].count(">"))
    if ids != GLUE_EXPECT["ids"]:
        bad.append("클론 순서 %s" % [x.split("_")[2] if "_" in x else x for x in ids])
    report.ok("F", F_NAMES[0], not bad,
              "param %s · design %s · rows %d · n_good %d · 중앙값 %d · fasta %d · 순서 c01~c04"
              % (GLUE_EXPECT["param_hash"], GLUE_EXPECT["design_hash"],
                 GLUE_EXPECT["merged_rows"], GLUE_EXPECT["n_good"],
                 GLUE_EXPECT["cdr3_median"], GLUE_EXPECT["fasta_n"]),
              " / ".join(bad))


def _check_f2(report, out):
    labels = [b["label"] for b in out["batches"]]
    cols = sorted(set(r[1] for r in out["merged"]["summary"]["rows"]))
    ok = len(set(labels)) == len(labels) and len(cols) == len(labels)
    report.ok("F", F_NAMES[1], ok,
              "같은 VH x 경쇄 카드 2 개가 서로 다른 label : " + " | ".join(labels),
              "label 이 겹칩니다 : %s (01 시트 batch 열 %s)" % (labels, cols))


def _check_f3(report, ns, core, dirs, cfg, meta):
    specs = [{"dir": dirs["b1"], "vh": "VH6", "chain": "kappa"},
             {"dir": dirs["b2"], "vh": "VH2", "chain": "lambda"}]
    out = _run_batches(ns, core, specs, cfg, meta)
    if not out.get("ok"):
        report.add("F", F_NAMES[2], FAIL, "실행 실패: " + "; ".join(out.get("errors", [])))
        return
    flags = {}
    for b in out["batches"]:
        for c in b["result"]["calls"]:
            flags[c["id"]] = c["flags"]
    f3 = flags.get(GLUE_EXPECT["ids"][2], [])
    f4 = flags.get(GLUE_EXPECT["ids"][3], [])
    d1, d2 = out["batches"][0]["design_hash"], out["batches"][1]["design_hash"]
    bad = []
    if "WRONG_CHAIN" not in f3:
        bad.append("c03 flags %s" % f3)
    if "WRONG_FAMILY" not in f4 or "WRONG_CHAIN" not in f4:
        bad.append("c04 flags %s" % f4)
    if d1 == d2:
        bad.append("design_hash 가 같음 %s" % d1)
    report.ok("F", F_NAMES[2], not bad,
              "c03 WRONG_CHAIN · c04 WRONG_FAMILY+WRONG_CHAIN · design_hash %s != %s"
              % (d1, d2), " / ".join(bad))


def _check_f4(report, ns, core, specs, cfg, meta):
    real = ns["core"]
    ns["core"] = _ShimCore(real)
    try:
        out = _run_batches(ns, core, specs, cfg, meta)
    finally:
        ns["core"] = real
    errs = out.get("errors") or []
    ok = (not out.get("ok")) and any("param_hash" in e for e in errs)
    report.ok("F", F_NAMES[3], ok,
              "두 번째 배치의 param_hash 를 바꾸면 ok=false · " +
              (errs[0][:40] if errs else ""),
              "ok=%s errors=%s (불일치를 통과시켰습니다)" % (out.get("ok"), errs))


def _check_glue_library(report, ns, core, cfg, meta):
    lib = dict(cfg)
    lib["analysis_mode"] = core.MODE_LIBRARY
    out = json.loads(ns["js_analyze"](json.dumps(lib), json.dumps(meta)))
    if not out.get("ok"):
        report.add("F", F_NAMES[4], FAIL, "실행 실패: " + "; ".join(out.get("errors", [])))
        return
    bad = []
    if out["config"]["design_hash"] != GLUE_EXPECT["lib_design_hash"]:
        bad.append("design_hash %s" % out["config"]["design_hash"])
    if out["composition"]["batch_vh_match"] is not None:
        bad.append("batch_vh_match %s" % out["composition"]["batch_vh_match"])
    if len(out["summary"]["rows"]) != GLUE_EXPECT["merged_rows"]:
        bad.append("rows %d" % len(out["summary"]["rows"]))
    report.ok("F", F_NAMES[4], not bad,
              "design_hash %s · batch_vh_match None · rows %d"
              % (GLUE_EXPECT["lib_design_hash"], GLUE_EXPECT["merged_rows"]),
              " / ".join(bad))


def _check_f6(report, box):
    sheets = box.get("sheets")
    if not sheets:
        report.add("F", F_NAMES[5], FAIL, "_write_output 이 호출되지 않았습니다")
        return
    titles = [s["title"] for s in sheets]
    bad = []
    if titles != GLUE_EXPECT["titles"]:
        bad.append("시트 제목 %s" % titles)
    n01 = len(sheets[0]["rows"]) if sheets else 0
    if n01 != GLUE_EXPECT["sheet01_rows"]:
        bad.append("01 행 %d" % n01)
    s04 = [s for s in sheets if s["title"].startswith("04")]
    marks = [r[1] for r in s04[0]["rows"] if r and r[0] == "배치"] if s04 else []
    if len(marks) != GLUE_EXPECT["block04"] or "전체 합계" not in marks:
        bad.append("04 블록 표시 %s" % marks)
    report.ok("F", F_NAMES[5], not bad,
              "7 시트 · 01 %d 행 · 04 블록 %d 개(배치 2 + 전체 1)"
              % (GLUE_EXPECT["sheet01_rows"], GLUE_EXPECT["block04"]),
              " / ".join(bad))


# =============================================================================
#  [G] 음성 대조군 실측 회귀
# =============================================================================
# testdata/negctrl/ 의 실측 .ab1 로 대조군 판정을 고정합니다. 디렉터리가 없으면
# [G] 전체를 건너뜁니다. 대조군 모드는 프라이머 판별을 쓰지 않으므로 FASTA 없이
# 돌립니다 (화면에서도 대조군 모드는 FASTA 를 요구하지 않습니다).
NEGCTRL_DIR = os.path.join(TESTDATA, "negctrl")
G_NAMES = ["G1 대조군 3 클론 판정", "G2 인서트 md5 고정",
           "G3 대조군 시트 구조", "G4 negctrl design_hash"]

_VEC = "%s-pAIM1-seq-For"       # 클론 ID = 확장자 뗀 파일명

NEGCTRL_EXPECT = {
    "verdicts": [(_VEC % "vec_1", "EMPTY_VECTOR"),
                 (_VEC % "vec_2", "CARRYOVER"),
                 (_VEC % "vec_3", "CARRYOVER")],
    # 실측해 박은 32 자 md5. 빈 문자열이면 아직 실측 전이라는 뜻이고,
    # 그때는 G2 가 측정값을 상세 칸에 찍고 실패합니다.
    # vec_1 은 NotI/AscI 가 모두 없어 인서트가 계산되지 않으므로 지문이 없습니다.
    "md5": {_VEC % "vec_2": "bdb1431f6378b361f1cc0f93cb172f38",   # 276 bp
            _VEC % "vec_3": "ed98c77f6c5403204bbd535d49c4abf1"},  # 297 bp
    # 라이브러리 c01 의 인서트 지문. vec_2 와 같아야 하며, 이 동일성이
    # 벡터 준비물 오염의 근거입니다.
    "c01_md5": "bdb1431f6378b361f1cc0f93cb172f38",
    "design_hash": "17a2b7ea",
    "titles": ["01_대조군판정", "02_구조QC상세", "03_대조군요약",
               "04_실행설정", "05_용어설명", "06_서열"],
    "sheet01_rows": 3,
}


def negctrl_inputs():
    d = os.path.join(ROOT, NEGCTRL_DIR)
    if not os.path.isdir(d):
        return None
    names = sorted(n for n in os.listdir(d) if n.lower().endswith(".ab1"))
    if not names:
        return None
    out = []
    for n in names:
        with open(os.path.join(d, n), "rb") as fh:
            out.append((n, fh.read()))
    return out


def check_negctrl_regression(report, core):
    files = negctrl_inputs()
    if files is None:
        for n in G_NAMES:
            report.add("G", n, SKIP, "testdata/negctrl/ 에 .ab1 이 없습니다")
        return

    cfg = core.build_config({"analysis_mode": core.MODE_NEGCTRL}, [])
    report.ok("G", G_NAMES[3],
              cfg["design_hash"] == NEGCTRL_EXPECT["design_hash"],
              NEGCTRL_EXPECT["design_hash"],
              "기대 %s 실제 %s" % (NEGCTRL_EXPECT["design_hash"], cfg["design_hash"]))

    meta = {"runtime": "verify.py " + VERIFY_VERSION, "timestamp": "0",
            "primer_file": "", "batch_label": "negctrl", "batch_date": ""}
    out = core.analyze(files, "", {"analysis_mode": core.MODE_NEGCTRL}, meta)
    if not out.get("ok") or not out.get("negctrl"):
        for n in G_NAMES[:3]:
            report.add("G", n, FAIL, "analyze 실패: " + "; ".join(out.get("errors", [])))
        return
    neg = out["negctrl"]
    by_id = dict((c["id"], c) for c in neg["clones"])

    # G1 판정과 근거
    bad = []
    for cid, want in NEGCTRL_EXPECT["verdicts"]:
        c = by_id.get(cid)
        if c is None:
            bad.append("%s 없음" % cid)
            continue
        if c["verdict"] != want:
            bad.append("%s %s(기대 %s)" % (cid.split("-")[0], c["verdict"], want))
        if "인서트 " not in c["reason"] or "프레임" not in c["reason"]:
            bad.append("%s reason 에 인서트/프레임 없음" % cid.split("-")[0])
    if len(neg["clones"]) != len(NEGCTRL_EXPECT["verdicts"]):
        bad.append("클론 %d 개 (기대 %d)"
                   % (len(neg["clones"]), len(NEGCTRL_EXPECT["verdicts"])))
    report.ok("G", G_NAMES[0], not bad,
              " · ".join("%s %s" % (c.split("-")[0], v)
                         for c, v in NEGCTRL_EXPECT["verdicts"]),
              " / ".join(bad))

    # G2 인서트 지문
    bad, seen = [], []
    for cid, want in sorted(NEGCTRL_EXPECT["md5"].items()):
        got = (by_id.get(cid) or {}).get("insert_md5", "")
        seen.append("%s %s" % (cid.split("-")[0], got or "(없음)"))
        if not want:
            bad.append("%s 기대 md5 미기입" % cid.split("-")[0])
        elif got != want:
            bad.append("%s md5 %s (기대 %s)" % (cid.split("-")[0], got, want))
    vec2 = (by_id.get(_VEC % "vec_2") or {}).get("insert_md5", "")
    if vec2 != NEGCTRL_EXPECT["c01_md5"]:
        bad.append("vec_2 md5 가 라이브러리 c01 (%s) 과 다름" % NEGCTRL_EXPECT["c01_md5"])
    report.ok("G", G_NAMES[1], not bad,
              "vec_2 md5 = 라이브러리 c01 (%s…) · 32 자 고정"
              % NEGCTRL_EXPECT["c01_md5"][:12],
              " / ".join(bad) + " · 실측 " + " / ".join(seen))

    # G3 시트 구조
    sheets = out["sheets"]
    bad = []
    titles = [s["title"] for s in sheets]
    if titles != NEGCTRL_EXPECT["titles"]:
        bad.append("시트 제목 %s" % titles)
    if sheets and len(sheets[0]["rows"]) != NEGCTRL_EXPECT["sheet01_rows"]:
        bad.append("01 행 %d" % len(sheets[0]["rows"]))
    s3 = [s for s in sheets if s["title"].startswith("03")]
    kinds = set(r[0] for r in s3[0]["rows"]) if s3 else set()
    for need in ("유형", "인서트 지문"):
        if need not in kinds:
            bad.append("03 에 '%s' 행 없음" % need)
    n_type = len([r for r in s3[0]["rows"] if r[0] == "유형"]) if s3 else 0
    if n_type != len(core.NEGCTRL_VERDICTS):
        bad.append("03 유형 행 %d (기대 %d)" % (n_type, len(core.NEGCTRL_VERDICTS)))
    report.ok("G", G_NAMES[2], not bad,
              "%d 시트 · 01 %d 행 · 03 에 유형 %d 종과 지문 목록"
              % (len(NEGCTRL_EXPECT["titles"]), NEGCTRL_EXPECT["sheet01_rows"],
                 len(core.NEGCTRL_VERDICTS)),
              " / ".join(bad))


# =============================================================================
#  실행
# =============================================================================
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    report = Report()
    html = read_text(HTML_FILE)
    blocks = script_blocks(html)
    js_all = "\n".join(blocks)
    glue = glue_source(js_all)
    js = js_all.replace(glue, "") if glue else js_all      # 파이썬 접착부 제외한 JS

    guard(report, "A", "AST 파싱", check_ast, report)
    guard(report, "A", "index.html JS 구문", check_js_syntax, report, html)
    guard(report, "A", "포맷 자리표시자 대조", check_format_arity, report)
    guard(report, "A", "함수 안 미정의 이름", check_undefined_names, report)

    sys.path.insert(0, ROOT)
    try:
        import core
    except Exception as e:
        core = None
        for sec in ("B", "C", "D"):
            report.add(sec, "core.py import", FAIL,
                       "core.py 를 불러올 수 없어 검사를 진행하지 못했습니다: %s: %s"
                       % (type(e).__name__, e))

    reg_out, reg_note, reg_status = None, "core.py 불러오기 실패", FAIL
    if core is not None:
        reg_out, reg_note, reg_status = run_analysis(core)

    if core is not None:
        core_tree = parse_py("core.py")
        if glue:
            gfn = find_function(ast.parse(glue), "js_docs")
            doc_keys, doc_sub = provided_keys(gfn) if gfn else (set(), {})
        else:
            doc_keys, doc_sub = set(), {}
        if not doc_keys:
            report.add("B", "js_docs 반환 키 추출", FAIL, "GLUE 의 js_docs 를 읽지 못했습니다")

        keyset = analyze_keyset(report, core, core_tree, reg_out, reg_note)

        guard(report, "B", "SUMMARY_HEADERS 대조", check_summary_headers, report, js, core)
        guard(report, "B", "JS 참조 키 대조", check_js_keys,
              report, js, core, keyset, doc_keys, doc_sub)
        guard(report, "B", "readConfig 가 DESIGN_KEYS 를 덮는가",
              check_readconfig_covers_design, report, js, core, glue or "")
        guard(report, "B", "CFG_DEFAULTS 대 CFG_DOC 키", check_cfg_doc, report, core)
        guard(report, "B", "DESIGN_DEFAULTS 대 DESIGN_DOC 키", check_design_doc, report, core)
        guard(report, "B", "analysis_mode 모드 값", check_analysis_mode, report, core)
        guard(report, "B", "대조군 verdict 처리", check_negctrl_vocab, report, core, html)
        guard(report, "B", "모호성 항목이 런타임 계산인가",
              check_ambiguity_runtime, report, core)
        guard(report, "B", "JUDGMENT_DESIGN_KEYS 가 DESIGN_DEFAULTS 에 존재",
              check_judgment_design_keys, report, core)
        guard(report, "B", "FLAG_SEV 대 qc_one 의 실제 플래그",
              check_flag_sev_keys, report, core, core_tree)
        guard(report, "B", "RULES 대 RULES_DOC 키", check_rules_doc, report, core)
        guard(report, "B", "옛 모듈 상수 잔존", check_moved_constants, report, core)
        guard(report, "B", "RULES 노출", check_rules_exposed,
              report, core, glue, reg_out)

        guard(report, "C", "index.html 6nt 이상 ACGT 리터럴",
              check_no_sequence, report, html)
        guard(report, "C", "CONST 서열이 index.html 에 등장",
              check_no_const_leak, report, html, core)
        guard(report, "C", "설정 라벨이 index.html 에 등장",
              check_no_label_leak, report, html, core)
        guard(report, "C", "모드 값이 index.html 에 등장",
              check_no_mode_leak, report, html, core)

        guard(report, "D", "회귀", check_regression,
              report, reg_out, reg_note, reg_status)

        guard(report, "E", "check_landmark 단위", check_landmark_units, report, core)
        guard(report, "E", "스터퍼 길이 판정 단위", check_stuffer_units, report, core)
        guard(report, "E", "대조군 판정 분기 단위", check_negctrl_units, report, core)
        guard(report, "E", "primer_ambiguity 단위", check_ambiguity_units, report, core)

    guard(report, "E", BADGE_LABEL, check_badge_states, report, js)

    if core is not None:
        guard(report, "F", "GLUE 경로", check_glue, report, glue, core)
        guard(report, "G", "대조군 실측 회귀", check_negctrl_regression, report, core)
    else:
        for n in F_NAMES:
            report.add("F", n, FAIL, "core.py 를 불러올 수 없어 GLUE 를 돌리지 못했습니다")
        for n in G_NAMES:
            report.add("G", n, FAIL, "core.py 를 불러올 수 없습니다")

    ver = "core %s / notebook %s" % (core.CORE_VERSION, core.NB_VERSION) \
        if core is not None else "core.py 불러오기 실패"
    print("pAIM1 scFv QC — verify.py %s   (%s)" % (VERIFY_VERSION, ver))
    print("")
    print_table(report)
    print("")
    parts = []
    for s in ("A", "B", "C", "D", "E", "F", "G"):
        parts.append("[%s] 통과 %d 실패 %d 건너뜀 %d"
                     % (s, report.count(s, PASS), report.count(s, FAIL),
                        report.count(s, SKIP)))
    print("   ".join(parts))
    return 1 if report.failed() else 0


if __name__ == "__main__":
    sys.exit(main())
